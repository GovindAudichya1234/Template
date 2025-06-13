
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.datavalidation import DataValidation
import os
import uuid

def apply_formulas_to_range(file_path, col_range, row_range, review_col, Rev,review_status_col):
    # Load workbook
    wb = load_workbook(file_path)
    sheet = wb.active

    # Parse column range and row range
    start_col, end_col = col_range.split('-')
    start_row, end_row = map(int, row_range.split('-'))

    start_col_index = column_index_from_string(start_col)
    end_col_index = column_index_from_string(end_col)

    # Delete data in the specified criteria column range from row 2 to end_row
    for col_idx in range(start_col_index, end_col_index + 1):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, end_row + 1):  # Deleting data from row 2 to end_row
            sheet[f"{col_letter}{row_idx}"].value = None

    # Define new headers to be written in row 2
    headers = [
        "Question Accuracy", "Question Distribution", "Answer Accuracy", "Answer Explanation",
        "Tagging bloom level", "Tagging complexity level", "Distractors","Finalised Question Status Count","Rejected Questions", "Learning Outcome",
        "No Repetition of PR Questions", "Topic Tagging", "Language and Grammar",
        "Copy Editing"
    ]
    
    # Write headers in row 2
    for col_idx, header in zip(range(start_col_index, end_col_index + 1), headers):
        sheet[f"{get_column_letter(col_idx)}2"].value = header
    
    saved_values = {"B": {}, "C": {},"H":{},"I":{}}  # Dictionary to store values for columns B and C
    if Rev in [ "R2", "R3"]:
        repowb = load_workbook(file_path,data_only=True)
        rub = repowb['AQR Rubrics']
        repo = repowb['Report Format']
        for col_idx in range(start_col_index, end_col_index + 1):
            col_letter = get_column_letter(col_idx)
            for row_idx in range(start_row, end_row + 1):
                cell = sheet[f"{col_letter}{row_idx}"]
                cell.value = None  # Clear the cell's value, removing formulas or text

        if Rev in ["R2", "R3"]:
            # Copy existing plain values from Column B
            for row_idx in range(start_row, end_row + 1):
                cell = repo[f"B{row_idx}"]
                saved_values["B"][row_idx] = cell.value
            for row_idx in range(start_row, end_row + 1):
                cell = rub[f"H{row_idx}"]
                saved_values["H"][row_idx] = cell.value  # Extract the plain value

        if Rev == "R3":
            # Copy existing plain values from Column C
            for row_idx in range(start_row, end_row + 1):
                cell = repo[f"B{row_idx}"]
                saved_values["B"][row_idx] = cell.value
            for row_idx in range(start_row, end_row + 1):
                cell = repo[f"C{row_idx}"]
                saved_values["C"][row_idx] = cell.value
            for row_idx in range(start_row, end_row + 1):
                cell = rub[f"H{row_idx}"]
                saved_values["H"][row_idx] = cell.value
            for row_idx in range(start_row, end_row + 1):
                cell = rub[f"I{row_idx}"]
                saved_values["I"][row_idx] = cell.value
        

        # ——————————————————————————————
    # Highlight & relabel rejected rows for R2/R3

    # ——————————————————————————————

    # Define formulas
    formulas = [
    f'=IF(SUM(ISNUMBER(SEARCH("Qs:", {review_col}ROW)) + ISNUMBER(SEARCH("QA:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("QD:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("AA:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("AE:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("Bloom:", {review_col}ROW)) + ISNUMBER(SEARCH("BT:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("Comp:", {review_col}ROW)) + ISNUMBER(SEARCH("CT:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("Dis:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(LOWER(TRIM({review_status_col}ROW)) = "closed", "Yes", "No")',
    f'=IF(ISNUMBER(SEARCH("Reject", {review_status_col}ROW)), "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("LO:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("Repq:", {review_col}ROW)) + ISNUMBER(SEARCH("QR:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("TopicT:", {review_col}ROW)) + ISNUMBER(SEARCH("TT:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("Lang:", {review_col}ROW)) + ISNUMBER(SEARCH("LG:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")',
    f'=IF(SUM(ISNUMBER(SEARCH("Ced:", {review_col}ROW)) + ISNUMBER(SEARCH("CE:", {review_col}ROW)) + ISNUMBER(SEARCH("Reject:", {review_col}ROW)))>0, "No", "Yes")'
    
]

    # Apply formulas to the specified range
    # Apply formulas to the specified range
    formula_index = 0
    for col_idx in range(start_col_index, end_col_index + 1):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(start_row, end_row + 1):
            template = formulas[formula_index % len(formulas)]
            filled   = (
                template
                .replace("ROW", str(row_idx))
                .replace("{review_col}", review_col)
                .replace("{review_status_col}", review_status_col)  # ← add this
            )
            sheet[f"{col_letter}{row_idx}"] = filled
        formula_index += 1



    # Add data validation for the specified range
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.prompt = "Please select Yes or No"
    dv.promptTitle = "Valid Options"

    # Add validation ranges explicitly
    for col_idx in range(start_col_index, end_col_index + 1):
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

    sheet.add_data_validation(dv)

    # Add COUNTIF formula in row 39 and percentage formula in row 40
    # Add COUNTIF and percentage formula one row after the end_row
    count_row = end_row + 1
    percentage_row = end_row + 2

    
    # Replace your old loop with this:
    for idx, col_idx in enumerate(range(start_col_index, end_col_index + 1)):
        col_letter = get_column_letter(col_idx)

        # For the 9th criterion (idx==8), count "No" (i.e. rejected) instead of "Yes"
        if idx == 8:
            countif_formula   = f'=COUNTIF({col_letter}{start_row}:{col_letter}{end_row}, "No")'
            percentage_formula = f'=({col_letter}{count_row}/{end_row - start_row + 1})*100'
        else:
            countif_formula   = f'=COUNTIF({col_letter}{start_row}:{col_letter}{end_row}, "Yes")'
            percentage_formula = f'=({col_letter}{count_row}/{end_row - start_row + 1})*100'

        sheet[f"{col_letter}{count_row}"]      = countif_formula
        sheet[f"{col_letter}{percentage_row}"] = percentage_formula

    from openpyxl.styles import PatternFill

    if Rev == "R2":
        fill = PatternFill("solid", fgColor="ADD8E6")   # light blue
        new_label = "Rejected R1"
    elif Rev == "R3":
        fill = PatternFill("solid", fgColor="FFCCCC")   # light red
        new_label = "Rejected R2"
    else:
        fill = None
        new_label = None

    if new_label and fill:
        max_col = sheet.max_column
        for r in range(start_row, end_row + 1):
            status_cell = sheet[f"{review_status_col}{r}"]
            text = str(status_cell.value or "")
            # already handled rows (Rejected R1 / Rejected R2) stay blank
            if text.strip() == "Rejected R1" and Rev == "R3":

                # blank criteria range
                for c in range(start_col_index, end_col_index + 1):
                    sheet.cell(row=r, column=c).value = None
            if "Reject" in text and not text.startswith("Rejected"):
                # blank criteria range
                for c in range(start_col_index, end_col_index + 1):
                    sheet.cell(row=r, column=c).value = None
                # colour full row
                for c in range(1, max_col + 1):
                    sheet.cell(row=r, column=c).fill = fill
                # overwrite status
                status_cell.value = new_label

    # Save workbook
    # Copy additional sheets from AQR file
    aqr_wb = load_workbook('AMT_AQR.xlsx')

    for sheet_name in ["AQR Rubrics", "Report Format"]:
        if sheet_name in aqr_wb.sheetnames:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            source_sheet = aqr_wb[sheet_name]
            target_sheet = wb.create_sheet(title=sheet_name)

            # Copy data and formatting manually
            for row in source_sheet.iter_rows():
                for cell in row:
                    target_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        from copy import copy

                        for row in source_sheet.iter_rows():
                            for cell in row:
                                tgt = target_sheet.cell(row=cell.row,
                                                        column=cell.column,
                                                        value=cell.value)
                                if cell.has_style:
                                    tgt.font          = copy(cell.font)
                                    tgt.border        = copy(cell.border)
                                    tgt.fill          = copy(cell.fill)
                                    tgt.number_format = copy(cell.number_format)
                                    tgt.protection    = copy(cell.protection)
                                    tgt.alignment     = copy(cell.alignment)

    def beautify_sheet(sheet, title_row=1):
        # Apply header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for cell in sheet[title_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = max_length + 2
    # Dynamically link percentages to "AQR Rubrics" and "Report Format" sheets
    # Dynamically link percentages to "AQR Rubrics" and "Report Format" sheets
        # Dynamically link percentages to "AQR Rubrics" and "Report Format" sheets
    beautify_sheet(wb["AQR Rubrics"])
    beautify_sheet(wb["Report Format"])
    rubrics_sheet = wb["AQR Rubrics"]
    report_sheet = wb["Report Format"]
    if Rev == "R1":
        RubSheetCol = 'H'
        RepSheetCol = 'B'
    elif Rev == "R2":
        RubSheetCol = 'I'
        RepSheetCol = 'C'
    else:
        RubSheetCol = 'J'
        RepSheetCol = 'D'

    if Rev == "R2":
        # Restore Column B values
        for row_idx, value in saved_values["B"].items():
            if value is not None:
                report_sheet[f"B{row_idx}"].value = value
        for row_idx, value in saved_values["H"].items():
            if value is not None:
                rubrics_sheet[f"H{row_idx}"].value = value

    if Rev == "R3":
        # Restore Column C values
        for row_idx, value in saved_values["B"].items():
            if value is not None:
                report_sheet[f"B{row_idx}"].value = value
        for row_idx, value in saved_values["C"].items():
            if value is not None:
                report_sheet[f"C{row_idx}"].value = value
        for row_idx, value in saved_values["H"].items():
            if value is not None:
                rubrics_sheet[f"H{row_idx}"].value = value
        for row_idx, value in saved_values["I"].items():
            if value is not None:
                rubrics_sheet[f"I{row_idx}"].value = value
    
    total_questions = end_row - start_row + 1
    current_row = 4
    formula_index = 0
    skip_rows = [13, 18]

    for col_idx in range(start_col_index, end_col_index + 1):
        col_letter = get_column_letter(col_idx)

        # 1) Skip your two blank rows
        while current_row in skip_rows:
            current_row += 1

        # 2) Pick the right formula for this header index
        if formula_index == 7:
            # Finalised Question Status Count → row 11
            formula = (
                f"=IFERROR("
                f"IF({sheet.title}!{col_letter}{percentage_row} > 94, 5, "
                f"IF({sheet.title}!{col_letter}{percentage_row} >= 86, 4, "
                f"IF({sheet.title}!{col_letter}{percentage_row} >= 84, 3, 1))), \"\")"
            )
        elif formula_index == 8:
            # Rejected Questions → row 12
            formula = (
                f"=IFERROR("
                f"IF({sheet.title}!{col_letter}{percentage_row} = 0, 5, "
                f"IF({sheet.title}!{col_letter}{percentage_row} <= 20, 2, 1)), \"\")"
            )
        else:
            # All the other 12 criteria
            formula = (
                f"=IFERROR("
                f"IF({sheet.title}!{col_letter}{percentage_row} <= 40, 1, "
                f"IF({sheet.title}!{col_letter}{percentage_row} <= 60, 2, "
                f"IF({sheet.title}!{col_letter}{percentage_row} <= 80, 3, "
                f"IF({sheet.title}!{col_letter}{percentage_row} <= 90, 4, 5)))), \"\")"
            )

        # 3) Write into both AQR Rubrics & Report Format
        rubrics_sheet[f"{RubSheetCol}{current_row}"] = formula
        report_sheet[f"{RepSheetCol}{current_row}"] = formula

        # advance to next
        current_row += 1
        formula_index += 1



    # Save workbook
    output_path = file_path.replace(".xlsx", "_processed.xlsx")
    wb.save(output_path)
    return output_path

# Streamlit app
st.title("Formula Application Tool with NEW AQR RUBRIC")
st.write("Instructions for using the tool:")
st.write("1. Upload an Excel file.")
st.write("2. Check if the provided AMT file have all the necessary columns like Question Accuracy ,Learning Outcome Accuracy ,No Repetition of PR Question ,Question Distribution ,Answer Accuracy ,Answer Explanation Accuracy ,Tagging bloom level ,Tagging complexity level ,Distractors ,Topic Tagging ,Language and Grammar ,Copy Editing")
st.write("3. If any column is missing add the column in the AMT file before uploading and it should be in the same order as mentioned above.")
st.write("4. Enter the column range and row range where you want to apply the formulas. (for e.g : Question Accuracy (AQ) to Copy Editing (BC) )")
st.write("5. Enter the row range where you want to apply the formulas. (for e.g : 3 to 38)")
st.write("6. Enter the column where the review specific comment is present. (for e.g : AK)")
st.write("7. For R2, R3 Add the R2 and R3 Comment Column before applying formula.")
st.write("8. If the formula is not getting applied or some issue happens. Delete the Yes or No from the criteria column and then do the process ")
st.write("9. Click on 'Apply Formula' button to apply the formulas.")


uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])
if uploaded_file:
    col_range = st.text_input("Enter Criteria Column Range (e.g., A-Z):")
    row_range = st.text_input("Enter Row Range (e.g., 3-38):")
    review_col = st.text_input("Enter Review Specific Comment Column (e.g., AK):")
    review_status_col = st.text_input("Enter Review Specific status Column (e.g., AK):")
    target_column_selection = st.selectbox("Select R1, R2, or R3:", ["R1", "R2", "R3"])
    if st.button("Apply Formula"):
        if col_range and row_range and review_col:
            # Save uploaded file to a temporary unique path
            unique_id = str(uuid.uuid4())
            temp_file_path = f"temp_{unique_id}.xlsx"
            with open(temp_file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())

            try:
                # Apply formulas and generate the output file
                output_path = apply_formulas_to_range(temp_file_path, col_range, row_range, review_col,target_column_selection,review_status_col)

                # Generate output file name using uploaded file name
                output_file_name = uploaded_file.name.replace(".xlsx", "_processed.xlsx")

                # Streamlit success message and download button
                st.success(f"Formulas applied successfully! Download the file below.")
                st.download_button(
                    label="Download Processed File",
                    data=open(output_path, "rb"),
                    file_name=output_file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Clean up temporary file
                os.remove(temp_file_path)

            except Exception as e:
                st.error(f"An error occurred: {e}")

        else:
            st.error("Please provide all inputs.")
